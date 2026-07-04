"""
3_Scripts/kml_matching.py

Step 2 of the pipeline.
Unzips NGO_1.kmz / NGO_2.kmz, parses every Placemark's polygon into a geodesic area +
centroid (shapely/pyproj, WGS84 ellipsoid), and matches each polygon to the Excel record
with the same UUID (loaded by data_cleaning.py). Validates:
  - Does every Excel record have a matching KML polygon (and vice versa)?
  - Is the KML polygon geometrically valid (no self-intersection)?
  - How far is the Kobo-recorded GPS point from the KML polygon's centroid?
  - Does the KML-measured area agree with the claimed area?

Outputs:
  2_Cleaned_Data/GIS_Validation_Report.xlsx
  <project_root>/_interim/merged.pkl   (handoff to overlap_analysis.py)

Run: python3 kml_matching.py   (after data_cleaning.py has been run at least once)
"""
import re
import zipfile
import warnings
from pathlib import Path

import pandas as pd
import numpy as np
from shapely.geometry import Polygon
from pyproj import Geod
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from data_cleaning import (RAW_DIR, CLEANED_DIR, INTERIM_DIR,
                            style_header, autosize, write_df, BODY_FONT, BORDER)

warnings.filterwarnings('ignore')
geod = Geod(ellps="WGS84")

NGO1_KMZ = RAW_DIR / 'NGO_1.kmz'
NGO2_KMZ = RAW_DIR / 'NGO_2.kmz'
NGO1_KML = INTERIM_DIR / 'ngo1_doc.kml'
NGO2_KML = INTERIM_DIR / 'ngo2_doc.kml'

DIST_THRESHOLD_M = 300      # GPS point vs KML centroid mismatch threshold
AREA_TOL_PCT = 15           # claimed vs KML-measured area mismatch threshold


# ---------------------------------------------------------------------------
# 1. UNZIP KMZ -> KML
# ---------------------------------------------------------------------------
def unzip_kmz(kmz_path, out_kml_path):
    with zipfile.ZipFile(kmz_path) as z:
        kml_name = next(n for n in z.namelist() if n.lower().endswith('.kml'))
        with z.open(kml_name) as src, open(out_kml_path, 'wb') as dst:
            dst.write(src.read())


# ---------------------------------------------------------------------------
# 2. KML PARSING
# ---------------------------------------------------------------------------
def parse_kml(path):
    content = open(path, encoding='utf-8').read()
    placemarks = re.findall(r'<Placemark>(.*?)</Placemark>', content, re.S)
    records = []
    for pm in placemarks:
        name_m = re.search(r'<name>(.*?)</name>', pm, re.S)
        raw_name = name_m.group(1).strip() if name_m else None
        uuid_clean = re.sub(r'[^0-9a-fA-F\-]', '', raw_name) if raw_name else None
        coord_m = re.search(r'<coordinates>(.*?)</coordinates>', pm, re.S)
        rec = {'raw_name': raw_name, 'kml_uuid': uuid_clean}
        if not coord_m:
            rec.update(area_acres=None, centroid_lat=None, centroid_lon=None,
                       num_points=0, valid_geom=False, polygon=None, geom_issue='No coordinates')
            records.append(rec)
            continue
        pts = []
        for tok in coord_m.group(1).strip().split():
            parts = tok.split(',')
            if len(parts) >= 2:
                try:
                    pts.append((float(parts[0]), float(parts[1])))
                except ValueError:
                    pass
        if len(pts) < 3:
            rec.update(area_acres=None, centroid_lat=None, centroid_lon=None,
                       num_points=len(pts), valid_geom=False, polygon=None, geom_issue='Fewer than 3 points')
            records.append(rec)
            continue
        try:
            poly = Polygon(pts)
            valid = poly.is_valid
            issue = None
            if not valid:
                issue = 'Self-intersecting polygon (auto-fixed for area calc)'
                poly_fixed = poly.buffer(0)
            else:
                poly_fixed = poly
            area_m2, _ = geod.geometry_area_perimeter(poly_fixed)
            centroid = poly_fixed.centroid
            rec.update(area_acres=abs(area_m2) / 4046.8564224, centroid_lat=centroid.y,
                       centroid_lon=centroid.x, num_points=len(pts), valid_geom=valid,
                       polygon=poly_fixed, geom_issue=issue)
        except Exception as e:
            rec.update(area_acres=None, centroid_lat=None, centroid_lon=None,
                       num_points=len(pts), valid_geom=False, polygon=None, geom_issue=f'Parse error: {e}')
        records.append(rec)
    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# 3. GPS-vs-KML-CENTROID DISTANCE
# ---------------------------------------------------------------------------
def gps_kml_distance(row):
    if pd.isna(row['kobo_lat']) or pd.isna(row['kobo_lon']) or pd.isna(row['kml_centroid_lat']):
        return np.nan
    _, _, dist = geod.inv(row['kobo_lon'], row['kobo_lat'], row['kml_centroid_lon'], row['kml_centroid_lat'])
    return dist


# ---------------------------------------------------------------------------
# 3b. SUSPICIOUS DUPLICATE-GPS CLUSTER DETECTION
#     Multiple *different* farmers recorded at the exact same coordinate (to ~1m
#     precision) is a strong signal of bulk-copy / lazy-pin-drop data collection,
#     not a legitimate coincidence. Records with GPS outside the valid project
#     region are excluded here since those are already caught as a GPS Issue —
#     mixing them in would dilute this specific, high-value finding.
# ---------------------------------------------------------------------------
SUSPICIOUS_MIN_RECORDS = 3
SUSPICIOUS_MIN_FARMERS = 2
COORD_PRECISION = 5  # ~1.1m at this latitude


def find_suspicious_gps_clusters(merged):
    d = merged[merged['kobo_lat'].between(10, 20) & merged['kobo_lon'].between(72, 80)].copy()
    d['rlat'] = d['kobo_lat'].round(COORD_PRECISION)
    d['rlon'] = d['kobo_lon'].round(COORD_PRECISION)
    rows = []
    suspicious_uuids = set()
    for (ngo, lat, lon), sub in d.groupby(['ngo', 'rlat', 'rlon']):
        if len(sub) < SUSPICIOUS_MIN_RECORDS:
            continue
        n_farmers = sub['farmer_name'].nunique()
        if n_farmers < SUSPICIOUS_MIN_FARMERS:
            continue
        rows.append({
            'ngo': ngo, 'lat': lat, 'lon': lon, 'records_at_location': len(sub),
            'distinct_farmers': n_farmers,
            'farmer_names': ', '.join(sorted(sub['farmer_name'].astype(str).unique())[:10]),
            'uuids': ', '.join(sub['uuid'].astype(str).tolist()),
            'block': sub['block_name_clean'].mode().iloc[0] if len(sub['block_name_clean'].mode()) else None,
        })
        suspicious_uuids.update(sub['uuid'].tolist())
    return pd.DataFrame(rows).sort_values('records_at_location', ascending=False), suspicious_uuids


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    master = pd.read_pickle(INTERIM_DIR / 'master.pkl')

    unzip_kmz(NGO1_KMZ, NGO1_KML)
    unzip_kmz(NGO2_KMZ, NGO2_KML)
    kml1 = parse_kml(NGO1_KML); kml1['ngo'] = 'NGO_1'
    kml2 = parse_kml(NGO2_KML); kml2['ngo'] = 'NGO_2'
    kml_all = pd.concat([kml1, kml2], ignore_index=True)
    print(f'Parsed {len(kml1)} NGO_1 polygons + {len(kml2)} NGO_2 polygons = {len(kml_all)} total')

    merged = master.merge(
        kml_all[['kml_uuid', 'raw_name', 'area_acres', 'centroid_lat', 'centroid_lon',
                 'num_points', 'valid_geom', 'geom_issue', 'ngo']],
        left_on=['uuid', 'ngo'], right_on=['kml_uuid', 'ngo'], how='left')
    merged.rename(columns={'area_acres': 'kml_area_acres', 'centroid_lat': 'kml_centroid_lat',
                            'centroid_lon': 'kml_centroid_lon'}, inplace=True)
    merged['has_kml_match'] = merged['kml_uuid'].notna()
    print('Records with a matching KML polygon:', merged['has_kml_match'].sum(), '/', len(merged))

    excel_uuids = set(master['uuid'])
    kml_uuids = set(kml_all['kml_uuid'].dropna())
    kml_only = kml_uuids - excel_uuids
    print('KML polygons with no matching Excel UUID:', len(kml_only))

    merged['gps_kml_distance_m'] = merged.apply(gps_kml_distance, axis=1)
    merged['gps_kml_mismatch'] = merged['gps_kml_distance_m'] > DIST_THRESHOLD_M
    print(f'GPS point >{DIST_THRESHOLD_M}m from KML centroid:', merged['gps_kml_mismatch'].sum())

    merged['area_diff_kml_vs_claimed'] = merged['kml_area_acres'] - merged['claimed_area']
    merged['area_diff_pct'] = (merged['area_diff_kml_vs_claimed'] / merged['claimed_area'] * 100).round(1)
    merged['area_mismatch_flag'] = merged['area_diff_pct'].abs() > AREA_TOL_PCT
    print(f'Area mismatch (>{AREA_TOL_PCT}%):', merged['area_mismatch_flag'].sum())
    print('Invalid/self-intersecting KML polygons:', merged['geom_issue'].notna().sum())

    # out-of-region GPS (column-shift errors) — carried forward for the final QA status
    bad_gps = (~merged['kobo_lat'].between(10, 20) | ~merged['kobo_lon'].between(72, 80)) & merged['kobo_lat'].notna()
    merged['kobo_gps_out_of_region'] = bad_gps
    print('GPS coordinates outside project region:', bad_gps.sum())

    def bad_year(x):
        return not bool(re.fullmatch(r'20\d\d', str(x).strip()))
    merged['year_planted_valid'] = ~merged['year_planted'].apply(bad_year)
    print('Malformed year_planted values:', (~merged['year_planted_valid']).sum())

    suspicious_df, suspicious_uuids = find_suspicious_gps_clusters(merged)
    merged['in_suspicious_gps_cluster'] = merged['uuid'].isin(suspicious_uuids)
    print(f'Suspicious duplicate-GPS clusters found: {len(suspicious_df)} '
          f'({merged["in_suspicious_gps_cluster"].sum()} records involved)')

    # ---------------- WRITE: GIS_Validation_Report.xlsx ----------------
    gis_cols = ['row_id', 'ngo', 'uuid', 'farmer_name', 'has_kml_match', 'geom_issue', 'num_points',
                'kobo_lat', 'kobo_lon', 'kml_centroid_lat', 'kml_centroid_lon', 'gps_kml_distance_m',
                'gps_kml_mismatch', 'kobo_gps_out_of_region', 'claimed_area', 'actual_area',
                'kml_area_acres', 'area_diff_pct', 'area_mismatch_flag', 'year_planted_valid']
    gis_df = merged[gis_cols].sort_values(['has_kml_match', 'gps_kml_mismatch', 'area_mismatch_flag'],
                                           ascending=[True, False, False])
    gis_df = gis_df.round({'gps_kml_distance_m': 0, 'kml_area_acres': 2, 'area_diff_pct': 1})

    wb = Workbook()
    ws = wb.active
    ws.title = 'GIS_Validation'
    write_df(ws, gis_df)

    kml_only_df = pd.DataFrame({'kml_uuid_with_no_excel_match': sorted(kml_only)})
    ws2 = wb.create_sheet('KML_Only_Placemarks')
    if len(kml_only_df):
        write_df(ws2, kml_only_df)
    else:
        ws2['A1'] = 'Every KML placemark UUID matches an Excel record.'

    ws3 = wb.create_sheet('Suspicious_GPS_Clusters')
    if len(suspicious_df):
        write_df(ws3, suspicious_df.drop(columns=['uuids']))
    else:
        ws3['A1'] = 'No suspicious duplicate-GPS clusters found (>=3 records, >=2 distinct farmers, same ~1m location).'
    wb.save(CLEANED_DIR / 'GIS_Validation_Report.xlsx')
    print('Saved 2_Cleaned_Data/GIS_Validation_Report.xlsx')

    # ---------------- Handoff to overlap_analysis.py ----------------
    kml_all.to_pickle(INTERIM_DIR / 'kml_all.pkl')
    merged.to_pickle(INTERIM_DIR / 'merged.pkl')
    print('Saved interim merged.pkl / kml_all.pkl for overlap_analysis.py')
    print('DONE: kml_matching.py')


if __name__ == '__main__':
    main()
