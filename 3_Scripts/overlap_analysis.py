"""
3_Scripts/overlap_analysis.py

Step 3 (final step) of the pipeline.
Detects spatially-overlapping KML plot polygons via a Shapely STRtree spatial index,
then rolls every check from data_cleaning.py + kml_matching.py + this script into a
3-tier severity status (Clean / Warning / Critical) and a set of operational reports:
per-NGO Data Reliability Score, top problematic villages, and a priority-actions list.

Outputs:
  2_Cleaned_Data/Cleaned_Master.xlsx
      README, Master_Cleaned, NGO_Summary, Reliability_Score,
      Top_Problem_Villages, Priority_Actions
  2_Cleaned_Data/GIS_Validation_Report.xlsx  (adds a "Plot_Overlaps" sheet, with the
      centroid coordinates of each overlapping pair so the dashboard can draw them)

Run: python3 overlap_analysis.py   (after data_cleaning.py and kml_matching.py)
"""
import warnings
from pathlib import Path

import pandas as pd
import numpy as np
from shapely.strtree import STRtree
from openpyxl import Workbook, load_workbook

from data_cleaning import CLEANED_DIR, INTERIM_DIR, write_df
from kml_matching import parse_kml, NGO1_KML, NGO2_KML

warnings.filterwarnings('ignore')

MIN_OVERLAP_PCT = 10        # detection threshold: flag a pair if overlap covers >= this % of either plot
CRITICAL_OVERLAP_PCT = 20   # a plot's own overlap share above this % is a Critical-severity issue


# ---------------------------------------------------------------------------
# 1. OVERLAP DETECTION (STRtree spatial index -> O(n log n) pairwise test)
# ---------------------------------------------------------------------------
def find_overlaps(df, ngo_label, min_overlap_pct=MIN_OVERLAP_PCT):
    df = df[df['polygon'].notna()].reset_index(drop=True)
    polys = df['polygon'].tolist()
    tree = STRtree(polys)
    seen = set()
    rows = []
    for i, poly in enumerate(polys):
        for j in tree.query(poly):
            j = int(j)
            if j <= i or (i, j) in seen:
                continue
            seen.add((i, j))
            other = polys[j]
            if not poly.intersects(other):
                continue
            inter = poly.intersection(other)
            if inter.is_empty or inter.area == 0:
                continue
            pct_i = inter.area / poly.area * 100 if poly.area > 0 else 0
            pct_j = inter.area / other.area * 100 if other.area > 0 else 0
            if max(pct_i, pct_j) < min_overlap_pct:
                continue
            rows.append({
                'ngo': ngo_label,
                'uuid_1': df.loc[i, 'kml_uuid'], 'uuid_2': df.loc[j, 'kml_uuid'],
                'overlap_pct_of_plot1': round(pct_i, 1), 'overlap_pct_of_plot2': round(pct_j, 1),
                'plot1_area_acres': round(df.loc[i, 'area_acres'], 3) if df.loc[i, 'area_acres'] else None,
                'plot2_area_acres': round(df.loc[j, 'area_acres'], 3) if df.loc[j, 'area_acres'] else None,
                'centroid1_lat': df.loc[i, 'centroid_lat'], 'centroid1_lon': df.loc[i, 'centroid_lon'],
                'centroid2_lat': df.loc[j, 'centroid_lat'], 'centroid2_lon': df.loc[j, 'centroid_lon'],
            })
    return pd.DataFrame(rows)


def max_overlap_pct_per_uuid(overlaps):
    """Each plot's own overlap share (not the other plot's), taking the worst pair it's in."""
    m = {}
    for _, r in overlaps.iterrows():
        m[r['uuid_1']] = max(m.get(r['uuid_1'], 0), r['overlap_pct_of_plot1'])
        m[r['uuid_2']] = max(m.get(r['uuid_2'], 0), r['overlap_pct_of_plot2'])
    return m


# ---------------------------------------------------------------------------
# 2. ISSUE CATEGORIES + 3-TIER SEVERITY
#    Every check from all three pipeline scripts collapses into a short list of
#    operational categories (what a field team would actually search/filter by),
#    plus one severity tier so the map and dashboard aren't just binary red/green.
# ---------------------------------------------------------------------------
CRITICAL_CATEGORIES = {
    'UUID Issue', 'Missing KML', 'GPS Issue', 'Geometry Overlap', 'Suspicious GPS Duplicate',
}
WARNING_CATEGORIES = {
    'Phone Issue', 'Area Mismatch', 'Name Standardization Needed',
    'Year Planted Issue', 'Geometry Self-Intersecting',
}


def classify_record(r, overlap_pct_map):
    categories = []

    # --- Critical-tier checks ---
    if not r['uuid_valid_format'] or r['is_duplicate_uuid']:
        categories.append('UUID Issue')
    if not r['has_kml_match']:
        categories.append('Missing KML')
    gps_issue = False
    if pd.isna(r['kobo_lat']) or pd.isna(r['kobo_lon']):
        gps_issue = True
    elif r['kobo_gps_out_of_region']:
        gps_issue = True
    elif pd.notna(r['gps_kml_mismatch']) and r['gps_kml_mismatch']:
        gps_issue = True
    if gps_issue:
        categories.append('GPS Issue')
    own_overlap = overlap_pct_map.get(r['uuid'], 0)
    if own_overlap >= CRITICAL_OVERLAP_PCT:
        categories.append('Geometry Overlap')
    if r.get('in_suspicious_gps_cluster', False):
        categories.append('Suspicious GPS Duplicate')

    # --- Warning-tier checks ---
    if r['phone_status'] != 'Valid' or r['is_duplicate_phone']:
        categories.append('Phone Issue')
    if pd.notna(r['area_mismatch_flag']) and r['area_mismatch_flag']:
        categories.append('Area Mismatch')
    if r['block_name_clean'] == 'INVALID/JUNK VALUE' or r['village_name_clean'] == 'INVALID/JUNK VALUE':
        categories.append('Name Standardization Needed')
    if not r['year_planted_valid']:
        categories.append('Year Planted Issue')
    if pd.notna(r['geom_issue']):
        categories.append('Geometry Self-Intersecting')
    if 0 < own_overlap < CRITICAL_OVERLAP_PCT:
        categories.append('Geometry Overlap (Minor)')

    if not categories:
        severity = 'Clean'
    elif any(c in CRITICAL_CATEGORIES for c in categories):
        severity = 'Critical'
    else:
        severity = 'Warning'

    return severity, '; '.join(categories) if categories else 'None'


# ---------------------------------------------------------------------------
# 3. NGO-WISE DATA RELIABILITY SCORE
#    100 minus weighted penalties for the checks that matter most operationally.
#    A simple, defensible, executive-readable summary metric.
# ---------------------------------------------------------------------------
RELIABILITY_WEIGHTS = {
    'invalid_gps_pct': 0.25,
    'duplicate_uuid_pct': 0.20,
    'missing_kml_pct': 0.20,
    'area_mismatch_pct': 0.15,
    'overlap_pct': 0.10,
    'invalid_phone_pct': 0.10,
}


def reliability_score(g, overlap_pct_map):
    invalid_gps = (g['kobo_lat'].isna() | g['kobo_lon'].isna() | g['kobo_gps_out_of_region']).mean() * 100
    dup_uuid = g['is_duplicate_uuid'].mean() * 100
    missing_kml = (~g['has_kml_match']).mean() * 100
    area_mismatch = g['area_mismatch_flag'].fillna(False).mean() * 100
    has_overlap = g['uuid'].map(lambda u: overlap_pct_map.get(u, 0) > 0).mean() * 100
    invalid_phone = (g['phone_status'] != 'Valid').mean() * 100

    penalty = (invalid_gps * RELIABILITY_WEIGHTS['invalid_gps_pct']
               + dup_uuid * RELIABILITY_WEIGHTS['duplicate_uuid_pct']
               + missing_kml * RELIABILITY_WEIGHTS['missing_kml_pct']
               + area_mismatch * RELIABILITY_WEIGHTS['area_mismatch_pct']
               + has_overlap * RELIABILITY_WEIGHTS['overlap_pct']
               + invalid_phone * RELIABILITY_WEIGHTS['invalid_phone_pct'])
    score = max(0, round(100 - penalty))
    return {
        'invalid_gps_pct': round(invalid_gps, 1), 'duplicate_uuid_pct': round(dup_uuid, 1),
        'missing_kml_pct': round(missing_kml, 1), 'area_mismatch_pct': round(area_mismatch, 1),
        'overlap_pct': round(has_overlap, 1), 'invalid_phone_pct': round(invalid_phone, 1),
        'reliability_score': score,
    }


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    merged = pd.read_pickle(INTERIM_DIR / 'merged.pkl')
    kml1 = parse_kml(NGO1_KML); kml1['ngo'] = 'NGO_1'
    kml2 = parse_kml(NGO2_KML); kml2['ngo'] = 'NGO_2'

    overlaps1 = find_overlaps(kml1, 'NGO_1')
    overlaps2 = find_overlaps(kml2, 'NGO_2')
    overlaps = pd.concat([overlaps1, overlaps2], ignore_index=True).sort_values(
        'overlap_pct_of_plot1', ascending=False)
    print('Overlapping plot pairs (>=%d%% overlap): %d' % (MIN_OVERLAP_PCT, len(overlaps)))
    overlap_pct_map = max_overlap_pct_per_uuid(overlaps)
    n_critical_overlap = sum(1 for v in overlap_pct_map.values() if v >= CRITICAL_OVERLAP_PCT)
    print(f'Plots with overlap >= {CRITICAL_OVERLAP_PCT}% (Critical tier): {n_critical_overlap}')

    # ---------------- Severity classification ----------------
    sev_cat = merged.apply(lambda r: classify_record(r, overlap_pct_map), axis=1)
    merged['severity'] = sev_cat.apply(lambda x: x[0])
    merged['issue_categories'] = sev_cat.apply(lambda x: x[1])
    merged['is_clean'] = merged['severity'] == 'Clean'
    merged['has_geometry_overlap'] = merged['uuid'].map(lambda u: overlap_pct_map.get(u, 0) > 0)
    merged['qa_status'] = merged['issue_categories'].where(merged['severity'] != 'Clean', 'Clean')

    print(merged['severity'].value_counts())
    print('Clean rate: {:.1f}%'.format(merged['is_clean'].mean() * 100))

    # ---------------- Issue-category breakdown (for the dashboard) ----------------
    all_cats = merged.loc[merged['severity'] != 'Clean', 'issue_categories'].str.split('; ').explode()
    category_counts = all_cats.value_counts().reset_index()
    category_counts.columns = ['issue_category', 'count']
    category_counts['pct_of_records'] = (category_counts['count'] / len(merged) * 100).round(1)

    # ---------------- Top problematic villages ----------------
    village_issues = merged[merged['severity'] != 'Clean'].groupby('village_name_clean')
    top_villages = village_issues.size().reset_index(name='issue_count').sort_values('issue_count', ascending=False)
    top_villages = top_villages[top_villages['village_name_clean'] != 'INVALID/JUNK VALUE'].head(20)

    def main_problem(village):
        cats = merged.loc[(merged['village_name_clean'] == village) & (merged['severity'] != 'Clean'),
                           'issue_categories'].str.split('; ').explode()
        return cats.value_counts().idxmax() if len(cats) else 'n/a'

    top_villages['main_problem'] = top_villages['village_name_clean'].apply(main_problem)
    top_villages['total_records'] = top_villages['village_name_clean'].apply(
        lambda v: (merged['village_name_clean'] == v).sum())
    top_villages['issue_rate_pct'] = (top_villages['issue_count'] / top_villages['total_records'] * 100).round(1)
    top_villages = top_villages.rename(columns={'village_name_clean': 'village'})[
        ['village', 'total_records', 'issue_count', 'issue_rate_pct', 'main_problem']]

    # ---------------- NGO-wise Data Reliability Score ----------------
    reliability_rows = []
    for ngo, g in merged.groupby('ngo'):
        rs = reliability_score(g, overlap_pct_map)
        rs['ngo'] = ngo
        rs['records'] = len(g)
        reliability_rows.append(rs)
    reliability_df = pd.DataFrame(reliability_rows)[
        ['ngo', 'records', 'reliability_score', 'invalid_gps_pct', 'duplicate_uuid_pct',
         'missing_kml_pct', 'area_mismatch_pct', 'overlap_pct', 'invalid_phone_pct']]
    print(reliability_df)

    # ---------------- Priority actions ----------------
    n_missing_kml = int((~merged['has_kml_match']).sum())
    n_overlap_pairs = len(overlaps)
    n_invalid_gps = int((merged['kobo_lat'].isna() | merged['kobo_lon'].isna()
                          | merged['kobo_gps_out_of_region']).sum())
    n_village_variants_before = merged['village_name_raw'].dropna().astype(str).str.strip().str.lower().nunique()
    n_dup_phone = int(merged['is_duplicate_phone'].sum())
    n_suspicious = int(merged.get('in_suspicious_gps_cluster', pd.Series(dtype=bool)).sum())
    n_dup_uuid = int(merged['is_duplicate_uuid'].sum())

    priority_actions = pd.DataFrame([
        {'priority': 1, 'action': f'Investigate {n_suspicious} records in suspicious duplicate-GPS clusters '
                                   '(same coordinate shared by multiple distinct farmers — likely bulk-copied pins)',
         'severity': 'Critical', 'records_affected': n_suspicious},
        {'priority': 2, 'action': f'Reverify {n_missing_kml} records with no matching KML polygon',
         'severity': 'Critical', 'records_affected': n_missing_kml},
        {'priority': 3, 'action': f'Investigate {n_overlap_pairs} polygon-overlap pairs '
                                   f'({n_critical_overlap} plots overlap >{CRITICAL_OVERLAP_PCT}%)',
         'severity': 'Critical', 'records_affected': n_overlap_pairs},
        {'priority': 4, 'action': f'Correct {n_invalid_gps} invalid/missing/out-of-region GPS points',
         'severity': 'Critical', 'records_affected': n_invalid_gps},
        {'priority': 5, 'action': f'Resolve {n_dup_uuid} duplicate UUID records',
         'severity': 'Critical', 'records_affected': n_dup_uuid},
        {'priority': 6, 'action': f'Standardize {n_village_variants_before} village-name spellings '
                                   '(canonical suggestions in Data_Quality_Report.xlsx)',
         'severity': 'Warning', 'records_affected': n_village_variants_before},
        {'priority': 7, 'action': f'Review {n_dup_phone} records sharing a phone number with a different farmer',
         'severity': 'Warning', 'records_affected': n_dup_phone},
    ])

    # ---------------- APPEND "Plot_Overlaps" sheet to GIS_Validation_Report.xlsx ----------------
    gis_path = CLEANED_DIR / 'GIS_Validation_Report.xlsx'
    wb = load_workbook(gis_path)
    if 'Plot_Overlaps' in wb.sheetnames:
        del wb['Plot_Overlaps']
    ws = wb.create_sheet('Plot_Overlaps')
    if len(overlaps):
        write_df(ws, overlaps)
    else:
        ws['A1'] = 'No overlapping plots found.'
    wb.save(gis_path)
    print('Updated 2_Cleaned_Data/GIS_Validation_Report.xlsx with Plot_Overlaps sheet')

    # ---------------- WRITE: Cleaned_Master.xlsx ----------------
    master_cols = ['row_id', 'ngo', 'uuid', 'uuid_valid_format', 'is_duplicate_uuid', 'farmer_name',
                   'surveyor_name', 'mobile_raw', 'phone_clean', 'phone_status', 'is_duplicate_phone',
                   'block_name_raw', 'block_name_clean', 'gp_name_raw', 'gp_name_clean',
                   'village_name_raw', 'village_name_clean', 'species', 'year_planted', 'year_planted_valid',
                   'kobo_lat', 'kobo_lon', 'kobo_gps_out_of_region', 'in_suspicious_gps_cluster',
                   'claimed_area', 'actual_area', 'kml_area_acres', 'area_diff_pct', 'area_mismatch_flag',
                   'has_kml_match', 'geom_issue', 'gps_kml_distance_m', 'gps_kml_mismatch',
                   'has_geometry_overlap', 'error_category', 'severity', 'issue_categories', 'qa_status',
                   'is_clean', 'submission_time', 'submitted_by']
    master_df = merged[master_cols].round(
        {'area_diff_pct': 1, 'gps_kml_distance_m': 0, 'kml_area_acres': 2})

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = 'README'
    ws2['B2'] = 'Makkala Vikasa (NGO_1) & YIDS (NGO_2) — Cleaned Master Dataset'
    ws2['B4'] = f'Total records: {len(merged)}'
    ws2['B5'] = f'Clean: {int((merged["severity"]=="Clean").sum())} | ' \
                f'Warning: {int((merged["severity"]=="Warning").sum())} | ' \
                f'Critical: {int((merged["severity"]=="Critical").sum())}'
    ws2['B6'] = f'Records with KML match: {int(merged["has_kml_match"].sum())} ' \
                f'({merged["has_kml_match"].mean()*100:.1f}%)'
    ws2['B7'] = f'Overlapping plot pairs: {len(overlaps)}'
    ws2['B8'] = f'Suspicious duplicate-GPS clusters: {len(merged.loc[merged.get("in_suspicious_gps_cluster", False)])} records'
    ws2['B10'] = 'See the Master_Cleaned sheet for the full row-level dataset (with a "severity" column: '
    ws2['B11'] = 'Clean / Warning / Critical, and an "issue_categories" column listing every check that failed).'
    ws2['B12'] = 'See Reliability_Score, Top_Problem_Villages and Priority_Actions for the operational summary.'
    ws3 = wb2.create_sheet('Master_Cleaned')
    write_df(ws3, master_df)

    ngo_rows = []
    for ngo, g in merged.groupby('ngo'):
        ngo_rows.append({
            'NGO': ngo, 'Total Records': len(g), 'Unique Farmers': g['farmer_name'].nunique(),
            'KML Match Rate %': round(g['has_kml_match'].mean() * 100, 1),
            'Clean %': round((g['severity'] == 'Clean').mean() * 100, 1),
            'Warning %': round((g['severity'] == 'Warning').mean() * 100, 1),
            'Critical %': round((g['severity'] == 'Critical').mean() * 100, 1),
            'Total Claimed Area (ac)': round(g['claimed_area'].sum(), 1),
            'Total KML Area (ac)': round(g['kml_area_acres'].sum(), 1),
        })
    ws4 = wb2.create_sheet('NGO_Summary')
    write_df(ws4, pd.DataFrame(ngo_rows))

    ws5 = wb2.create_sheet('Reliability_Score')
    write_df(ws5, reliability_df)

    ws6 = wb2.create_sheet('Issue_Category_Breakdown')
    write_df(ws6, category_counts)

    ws7 = wb2.create_sheet('Top_Problem_Villages')
    write_df(ws7, top_villages)

    ws8 = wb2.create_sheet('Priority_Actions')
    write_df(ws8, priority_actions)

    wb2.save(CLEANED_DIR / 'Cleaned_Master.xlsx')
    print('Saved 2_Cleaned_Data/Cleaned_Master.xlsx')

    # ---------------- Final handoff (for the dashboard, if it wants raw data) ----------------
    merged.to_pickle(INTERIM_DIR / 'final.pkl')
    overlaps.to_pickle(INTERIM_DIR / 'overlaps.pkl')
    print('DONE: overlap_analysis.py')


if __name__ == '__main__':
    main()
