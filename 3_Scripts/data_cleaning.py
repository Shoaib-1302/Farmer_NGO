"""
3_Scripts/data_cleaning.py

Step 1 of the pipeline.
Loads NGO_1.xlsx and NGO_2.xlsx (Kobo Collect exports), unifies them into one schema,
and runs all Excel-level data quality checks:
  - UUID format / duplicate detection
  - Phone number validation & duplicate detection
  - Block / GP / Village name standardization (fuzzy clustering of spelling variants)
  - Missing-value audit
  - Standardization of the free-text "Error" notes already logged by the NGOs

Outputs:
  2_Cleaned_Data/Duplicate_Records.xlsx
  2_Cleaned_Data/Data_Quality_Report.xlsx
  <project_root>/_interim/master.pkl   (handoff to kml_matching.py)

Run: python3 data_cleaning.py   (from inside 3_Scripts, or any cwd — paths are resolved
relative to this file)
"""
import re
import warnings
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ---------------------------------------------------------------------------
# PATHS
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / '1_Raw_Data'
CLEANED_DIR = PROJECT_ROOT / '2_Cleaned_Data'
INTERIM_DIR = PROJECT_ROOT / '_interim'
CLEANED_DIR.mkdir(exist_ok=True)
INTERIM_DIR.mkdir(exist_ok=True)

NGO1_XLSX = RAW_DIR / 'NGO_1.xlsx'   # Makkala Vikasa
NGO2_XLSX = RAW_DIR / 'NGO_2.xlsx'   # YIDS

# ---------------------------------------------------------------------------
# VALIDATION PATTERNS
# ---------------------------------------------------------------------------
UUID_RE = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$')
PHONE_RE = re.compile(r'^[6-9][0-9]{9}$')


# ---------------------------------------------------------------------------
# 1. EXCEL LOADERS — map each NGO's raw column names into one common schema
# ---------------------------------------------------------------------------
def load_ngo1(path):
    """Makkala Vikasa Kobo export."""
    df = pd.read_excel(path, sheet_name='Sheet1')
    out = pd.DataFrame(index=df.index)
    out['ngo'] = 'NGO_1'
    out['uuid'] = df['_uuid'].astype(str).str.strip()
    out['surveyor_name'] = df['Surveryor_s_Name']
    out['farmer_name'] = df['Farmer_s_Name']
    out['mobile_raw'] = df['Farmer_s_Mobile_Number']
    out['block_name_raw'] = df['Block_Name']
    out['gp_name_raw'] = df['Gram_Panachayat_Name']
    out['village_name_raw'] = df['Village_Name']
    out['species'] = df['tree_cat']
    out['year_planted'] = df['Year_planted']
    out['kobo_lat'] = pd.to_numeric(df['_Record_your_current_location_latitude'], errors='coerce')
    out['kobo_lon'] = pd.to_numeric(df['_Record_your_current_location_longitude'], errors='coerce')
    out['claimed_area'] = pd.to_numeric(df['Claimed area'], errors='coerce')
    out['actual_area'] = pd.to_numeric(df['Actual area'], errors='coerce')
    out['trees_per_acre'] = pd.to_numeric(df['Trees/acre'], errors='coerce')
    out['existing_error_note'] = df['Error']
    out['submission_time'] = df['_submission_time']
    out['submitted_by'] = df['_submitted_by']
    return out


def load_ngo2(path):
    """YIDS Kobo export."""
    df = pd.read_excel(path, sheet_name='Sheet1')
    out = pd.DataFrame(index=df.index)
    out['ngo'] = 'NGO_2'
    out['uuid'] = df['_uuid'].astype(str).str.strip()
    out['surveyor_name'] = df['Surveyor_name']
    out['farmer_name'] = df['Farmer_name']
    out['mobile_raw'] = df["Farmer's_Mobile_number"]
    out['block_name_raw'] = df['Block_name']
    out['gp_name_raw'] = df['GP_Name']
    out['village_name_raw'] = df['Village_Name']
    out['species'] = df['Participant categories (pick all that apply)']
    out['year_planted'] = df['Year_planted']
    out['kobo_lat'] = pd.to_numeric(df['_Record your current location_latitude'], errors='coerce')
    out['kobo_lon'] = pd.to_numeric(df['_Record your current location_longitude'], errors='coerce')
    out['claimed_area'] = pd.to_numeric(df['Claimed area (acre)'], errors='coerce')
    out['actual_area'] = pd.to_numeric(df['Actual area (acrea)'], errors='coerce')
    out['trees_per_acre'] = np.nan
    out['existing_error_note'] = df['Error']
    out['submission_time'] = df['_submission_time']
    out['submitted_by'] = df['_submitted_by']
    return out


# ---------------------------------------------------------------------------
# 2. PHONE VALIDATION
# ---------------------------------------------------------------------------
def validate_phone(x):
    if pd.isna(x):
        return 'Missing'
    s = re.sub(r'\D', '', str(x))
    if s.startswith('91') and len(s) == 12:
        s = s[2:]
    if len(s) != 10:
        return f'Invalid length ({len(s)} digits)'
    if not PHONE_RE.match(s):
        return 'Invalid format (does not start with 6-9)'
    return 'Valid'


def clean_phone(x):
    if pd.isna(x):
        return None
    s = re.sub(r'\D', '', str(x))
    if s.startswith('91') and len(s) == 12:
        s = s[2:]
    return s


# ---------------------------------------------------------------------------
# 3. NAME STANDARDIZATION (fuzzy clustering of village/block/GP names)
# ---------------------------------------------------------------------------
def normalize_basic(s):
    if pd.isna(s):
        return s
    s = re.sub(r'\s+', ' ', str(s).strip().lower())
    return s


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def flag_junk_name(s):
    if pd.isna(s):
        return True
    s = str(s).strip()
    if s in ('', '0'):
        return True
    if re.fullmatch(r'[0-9\?\.\-]+', s):
        return True
    if len(s) <= 1:
        return True
    return False


def build_name_correction_map(series, threshold=0.84):
    """Cluster similar strings and map each raw value -> canonical (most frequent) form."""
    vals = series.dropna().astype(str)
    norm = vals.apply(normalize_basic)
    counts = norm.value_counts()
    uniques = list(counts.index)
    clusters = []
    assigned = {}
    for u in uniques:
        placed = False
        for ci, cluster in enumerate(clusters):
            if similar(u, cluster[0]) >= threshold:
                clusters[ci].append(u)
                assigned[u] = ci
                placed = True
                break
        if not placed:
            clusters.append([u])
            assigned[u] = len(clusters) - 1
    canonical = {ci: max(cluster, key=lambda x: counts[x]).title() for ci, cluster in enumerate(clusters)}
    corr_map = {u: canonical[assigned[u]] for u in uniques}
    return norm.map(corr_map), corr_map


def clean_and_map(series, threshold=0.84):
    junk_mask = series.apply(flag_junk_name)
    mapped, corr_map = build_name_correction_map(series[~junk_mask], threshold=threshold)
    full = pd.Series(index=series.index, dtype=object)
    full[~junk_mask] = mapped
    full[junk_mask] = 'INVALID/JUNK VALUE'
    return full, corr_map


def corr_table(raw_series, corr_map, field):
    rows = []
    norm = raw_series.dropna().astype(str).apply(normalize_basic)
    vc = norm.value_counts()
    for raw_norm, cnt in vc.items():
        canon = corr_map.get(raw_norm)
        if canon and canon.lower() != raw_norm:
            rows.append({'field': field, 'raw_value': raw_norm, 'occurrences': cnt, 'suggested_correction': canon})
    return pd.DataFrame(rows).sort_values(['field', 'occurrences'], ascending=[True, False])


# ---------------------------------------------------------------------------
# 4. STANDARDIZE FREE-TEXT "ERROR" NOTES INTO CATEGORIES
# ---------------------------------------------------------------------------
ERROR_CATEGORY_RULES = [
    ('Correct / No Issue', [r'\bcorrect\b', r'^ok$', r'^good$']),
    ('Area Mismatch', [r'area', r'\bacre\b']),
    ('Missing/Incorrect GPS', [r'lat', r'\blon', r'gps', r'location', r'no\s*point', r'wrong point']),
    ('Image Quality Issue', [r'image', r'photo', r'video', r'black']),
    ('Phone Number Issue', [r'phone', r'mobile', r'contact']),
    ('Plantation Year Issue', [r'plantation', r'year']),
    ('Boundary/Border Issue', [r'border', r'boundary', r'overlap']),
    ('Vegetation/Species Issue', [r'vegetation', r'species', r'tree']),
    ('Duplicate Record', [r'duplicate', r'same uuid', r'repeat']),
]


def categorize_error_note(note):
    if pd.isna(note) or str(note).strip() == '':
        return 'No Note / Not Reviewed'
    text = str(note).lower()
    if UUID_RE.match(text.strip()):
        return 'Malformed Note (UUID pasted into Error field)'
    matched = [cat for cat, patterns in ERROR_CATEGORY_RULES if any(re.search(p, text) for p in patterns)]
    if not matched:
        return 'Other / Uncategorized'
    if 'Correct / No Issue' in matched and len(matched) > 1:
        matched.remove('Correct / No Issue')
    return '; '.join(matched)


# ---------------------------------------------------------------------------
# EXCEL WRITE HELPERS
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill('solid', start_color='1F4E3D', end_color='1F4E3D')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
BODY_FONT = Font(name='Arial', size=10)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, df, max_width=40):
    for i, col in enumerate(df.columns):
        letter = get_column_letter(i + 1)
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:500]])
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), max_width)


def write_df(ws, df, start_row=1):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    for i, (_, r) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if isinstance(v, np.integer):
                v = int(v)
            elif isinstance(v, np.floating):
                v = float(v) if not np.isnan(v) else None
            elif pd.isna(v):
                v = None
            cell = ws.cell(row=i, column=j, value=v)
            cell.font, cell.border = BODY_FONT, BORDER
    style_header(ws, row=start_row, ncols=len(df.columns))
    autosize(ws, df)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    ngo1 = load_ngo1(NGO1_XLSX)
    ngo2 = load_ngo2(NGO2_XLSX)
    master = pd.concat([ngo1, ngo2], ignore_index=True)
    master['row_id'] = master.index + 1
    print(f'Loaded {len(ngo1)} NGO_1 records + {len(ngo2)} NGO_2 records = {len(master)} total')

    # --- UUID checks ---
    master['uuid_valid_format'] = master['uuid'].apply(lambda x: bool(UUID_RE.match(str(x))))
    uuid_counts = master['uuid'].value_counts()
    master['is_duplicate_uuid'] = master['uuid'].map(uuid_counts) > 1
    print('Invalid-format UUIDs:', (~master['uuid_valid_format']).sum())
    print('Duplicate UUID rows:', master['is_duplicate_uuid'].sum())

    # --- Phone checks ---
    master['phone_status'] = master['mobile_raw'].apply(validate_phone)
    master['phone_clean'] = master['mobile_raw'].apply(clean_phone)
    phone_counts = master.loc[master['phone_status'] == 'Valid', 'phone_clean'].value_counts()
    dup_phones = phone_counts[phone_counts > 1]
    master['is_duplicate_phone'] = master['phone_clean'].isin(dup_phones.index) & (master['phone_status'] == 'Valid')
    print('Invalid phone records:', (master['phone_status'] != 'Valid').sum())
    print('Duplicate-phone records:', master['is_duplicate_phone'].sum())

    # --- Name standardization ---
    master['block_name_clean'], block_map = clean_and_map(master['block_name_raw'])
    master['village_name_clean'], village_map = clean_and_map(master['village_name_raw'])
    master['gp_name_clean'], gp_map = clean_and_map(master['gp_name_raw'])
    name_corrections = pd.concat([
        corr_table(master['block_name_raw'], block_map, 'Block_Name'),
        corr_table(master['village_name_raw'], village_map, 'Village_Name'),
        corr_table(master['gp_name_raw'], gp_map, 'GP_Name'),
    ], ignore_index=True)
    print('Suggested name auto-corrections:', len(name_corrections))

    def n_raw_variants(series):
        return series.dropna().astype(str).str.strip().str.lower().nunique()

    def n_clean_variants(clean_series):
        return clean_series[clean_series != 'INVALID/JUNK VALUE'].nunique()

    impact_rows = [
        {'metric': 'Block name spelling variants', 'raw': n_raw_variants(master['block_name_raw']),
         'after_standardization': n_clean_variants(master['block_name_clean'])},
        {'metric': 'Village name spelling variants', 'raw': n_raw_variants(master['village_name_raw']),
         'after_standardization': n_clean_variants(master['village_name_clean'])},
        {'metric': 'GP name spelling variants', 'raw': n_raw_variants(master['gp_name_raw']),
         'after_standardization': n_clean_variants(master['gp_name_clean'])},
    ]
    print('Standardization impact:', impact_rows)

    # --- Missing value audit ---
    critical_fields = ['farmer_name', 'mobile_raw', 'block_name_raw', 'village_name_raw', 'species',
                        'kobo_lat', 'kobo_lon']
    missing_report = master[critical_fields].isna().sum().rename('missing_count').to_frame()
    missing_report['missing_pct'] = (missing_report['missing_count'] / len(master) * 100).round(2)
    missing_report = missing_report.reset_index().rename(columns={'index': 'field'})

    # --- Error-note categorization ---
    master['error_category'] = master['existing_error_note'].apply(categorize_error_note)
    error_cat_counts = master['error_category'].value_counts().reset_index()
    error_cat_counts.columns = ['error_category', 'count']
    error_cat_counts['pct_of_records'] = (error_cat_counts['count'] / len(master) * 100).round(1)

    n_raw_note_phrasings = master['existing_error_note'].dropna().astype(str).str.strip().str.lower().nunique()
    n_standard_categories = master.loc[master['error_category'] != 'No Note / Not Reviewed', 'error_category'].nunique()
    impact_rows.append({'metric': 'Free-text "Error" note phrasings (partner logs)',
                         'raw': n_raw_note_phrasings, 'after_standardization': n_standard_categories})
    n_dup_uuid_rows = int(master['is_duplicate_uuid'].sum())
    n_dup_uuid_groups = master.loc[master['is_duplicate_uuid'], 'uuid'].nunique()
    impact_rows.append({'metric': 'Duplicate UUID rows (in N groups)',
                         'raw': f'{n_dup_uuid_rows} rows in {n_dup_uuid_groups} groups',
                         'after_standardization': f'{n_dup_uuid_groups} unique plots if 1 kept per group'})
    standardization_impact = pd.DataFrame(impact_rows)
    standardization_impact['raw'] = standardization_impact['raw'].astype(str)
    standardization_impact['after_standardization'] = standardization_impact['after_standardization'].astype(str)

    # --- Duplicate reports ---
    dup_uuid_report = master[master['is_duplicate_uuid']].sort_values('uuid')[
        ['row_id', 'ngo', 'uuid', 'farmer_name', 'mobile_raw', 'block_name_raw', 'village_name_raw',
         'submission_time']]
    dup_phone_report = master.loc[master['is_duplicate_phone'],
        ['row_id', 'ngo', 'uuid', 'farmer_name', 'phone_clean', 'block_name_raw', 'village_name_raw',
         'submission_time']].sort_values('phone_clean')
    if len(dup_phone_report):
        dup_phone_report['same_phone_group_size'] = dup_phone_report.groupby('phone_clean')['phone_clean'] \
            .transform('size')

    # ---------------- WRITE: Duplicate_Records.xlsx ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = 'Duplicate_UUIDs'
    if len(dup_uuid_report):
        write_df(ws, dup_uuid_report)
    else:
        ws['A1'] = 'No duplicate UUIDs found.'
    ws2 = wb.create_sheet('Duplicate_Phones')
    if len(dup_phone_report):
        write_df(ws2, dup_phone_report)
    else:
        ws2['A1'] = 'No duplicate phone numbers found.'
    wb.save(CLEANED_DIR / 'Duplicate_Records.xlsx')
    print('Saved 2_Cleaned_Data/Duplicate_Records.xlsx')

    # ---------------- WRITE: Data_Quality_Report.xlsx ----------------
    wb2 = Workbook()
    ws3 = wb2.active
    ws3.title = 'Missing_Values'
    write_df(ws3, missing_report)
    ws4 = wb2.create_sheet('Name_Corrections')
    write_df(ws4, name_corrections)
    ws5 = wb2.create_sheet('Error_Categories')
    write_df(ws5, error_cat_counts)
    ws6 = wb2.create_sheet('Phone_Validation_Summary')
    phone_summary = master['phone_status'].value_counts().reset_index()
    phone_summary.columns = ['phone_status', 'count']
    write_df(ws6, phone_summary)
    ws7 = wb2.create_sheet('Standardization_Impact')
    write_df(ws7, standardization_impact)
    wb2.save(CLEANED_DIR / 'Data_Quality_Report.xlsx')
    print('Saved 2_Cleaned_Data/Data_Quality_Report.xlsx')

    # ---------------- Handoff to kml_matching.py ----------------
    master.to_pickle(INTERIM_DIR / 'master.pkl')
    print('Saved interim master.pkl for kml_matching.py')
    print('DONE: data_cleaning.py')


if __name__ == '__main__':
    main()
